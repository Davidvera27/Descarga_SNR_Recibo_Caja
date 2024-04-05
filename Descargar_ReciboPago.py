import sys
from selenium.webdriver.common.keys import Keys
from PyQt5.QtCore import QUrl
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QHeaderView, QPushButton, QLabel, QLineEdit, QHBoxLayout
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Resultados de la Consulta")
        self.setGeometry(100, 100, 800, 600)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(3)
        self.table_widget.setHorizontalHeaderLabels(["ESCRITURA", "Estado del Recibo", "ABRIR CASO"])
        self.layout.addWidget(self.table_widget)
        self.central_widget.setLayout(self.layout)
        self.headless = True  # Por defecto, el navegador se ejecutará en modo Headless

        options = Options()
        options.add_argument("--start-maximized")
        options.add_argument("--headless")  # Agregar esta línea para ejecutar en modo Headless
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)

        # Realizar la consulta inicial sobre todos los casos
        self.consultar_casos()
        
        # Cargar estilos CSS personalizados
        with open("PYTHON/WebScraping/Excel/Descarga_SNR_Recibo_Caja/styles_Interfaz_ReciboPago.css", "r") as f:
            self.setStyleSheet(f.read())

        # Mostrar la interfaz con los resultados de la consulta
        self.mostrar_interfaz()

    def consultar_casos(self):
        try:
            # Cargar el archivo Excel y seleccionar la hoja principal
            wb = load_workbook("C:/Users/DAVID/Desktop/DAVID/N-15/DAVID/LIBROS XLSM/HISTORICO.xlsm", data_only=True)
            ws = wb["PRINCIPAL"]

            # Iterar sobre las filas y procesar aquellas que cumplan con las condiciones
            for row in ws.iter_rows(min_row=2, max_col=11):
                try:
                    if verificar_condiciones(row):
                        nir = row[3].value
                        escritura_numero = row[1].value
                        self.driver.get("https://radicacion.supernotariado.gov.co/app/inicio.dma")
                        # Seleccionar la opción "Pagos en linea"
                        self.driver.find_element(By.ID, "formLinks:paymentLink").click()

                        # Esperar a que aparezca el campo de búsqueda NIR
                        WebDriverWait(self.driver, 10).until(
                            EC.visibility_of_element_located((By.ID, "filtroForm:j_idt41"))
                        )
                        estado_recibo = consultar_nir(self.driver, nir, escritura_numero)
                        self.mostrar_resultado(escritura_numero, estado_recibo, nir)
                except TimeoutException as e:
                    print(f"Tiempo de espera agotado: {str(e)}")
                except WebDriverException as e:
                    print(f"Error de WebDriver: {str(e)}")
        except Exception as e:
            print(f"Error al consultar casos: {str(e)}")

    def mostrar_interfaz(self):
        self.show()

    def mostrar_resultado(self, escritura_numero, estado_recibo, nir):
        # Obtener el número de filas actual en la tabla
        row_position = self.table_widget.rowCount()
        self.table_widget.insertRow(row_position)
        self.table_widget.setItem(row_position, 0, QTableWidgetItem(str(escritura_numero)))
        self.table_widget.setItem(row_position, 1, QTableWidgetItem(estado_recibo))

        # Crear el botón "ABRIR" y conectarlo con la función abrir_caso
        abrir_button = QPushButton("ABRIR", self)
        abrir_button.clicked.connect(lambda _, nir_value=nir: self.abrir_caso(nir_value))
        self.table_widget.setCellWidget(row_position, 2, abrir_button)
        abrir_button.clicked.connect(self.mostrar_navegador_normal)
        
    def mostrar_navegador_normal(self):
        # Esta función se llama cuando se presiona el botón "ABRIR"
        # Cambiar self.headless a False para ejecutar el navegador en modo normal
        self.headless = False        

    def abrir_caso(self, nir):
        try:
            # Abrir el navegador en modo Headless solo si self.headless es True
            if self.headless:
                # Configuración de Selenium en modo Headless
                options = Options()
                options.add_argument("--start-maximized")                
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
            else:
                # Configuración de Selenium en modo normal
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service)

            # Navegar a la URL del sitio web
            self.driver.get("https://radicacion.supernotariado.gov.co/app/inicio.dma")

            # Seleccionar la opción "Pagos en línea"
            self.driver.find_element(By.ID, "formLinks:paymentLink").click()

            # Esperar a que aparezca el campo de búsqueda NIR
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, "filtroForm:j_idt41")))

            # Ingresar el NIR en el campo de texto
            input_nir = self.driver.find_element(By.ID, "filtroForm:j_idt41")
            input_nir.clear()
            input_nir.send_keys(nir)
            input_nir.send_keys(Keys.RETURN)  # Presionar Enter para buscar el NIR

        except Exception as e:
            print(f"Error al abrir el caso: {str(e)}")

    def closeEvent(self, event):
        # Cerrar el navegador al cerrar la ventana principal
        if self.driver is not None:
            self.driver.quit()
        event.accept()  # Aceptar el evento de cierre de la ventana principal

    def busqueda_avanzada(self):
        self.busqueda_avanzada_window = QWidget()
        self.busqueda_avanzada_window.setWindowTitle("Búsqueda Avanzada")
        self.busqueda_avanzada_window.setGeometry(200, 200, 400, 200)

        layout = QVBoxLayout()

        self.label_escritura = QLabel("Número de Escritura:")
        self.line_edit_escritura = QLineEdit()
        self.label_nir = QLabel("NIR:")
        self.line_edit_nir = QLineEdit()

        layout.addWidget(self.label_escritura)
        layout.addWidget(self.line_edit_escritura)
        layout.addWidget(self.label_nir)
        layout.addWidget(self.line_edit_nir)

        search_button = QPushButton("BUSCAR")
        search_button.clicked.connect(self.buscar_caso)
        layout.addWidget(search_button)

        open_button = QPushButton("ABRIR")
        open_button.clicked.connect(self.abrir_caso)
        layout.addWidget(open_button)

        self.busqueda_avanzada_window.setLayout(layout)
        self.busqueda_avanzada_window.show()

    def buscar_caso(self):
        escritura = self.line_edit_escritura.text().strip()
        nir = self.line_edit_nir.text().strip()
        
        # Cargar el archivo Excel y seleccionar la hoja principal
        wb = load_workbook("C:/Users/DAVID/Desktop/DAVID/N-15/DAVID/LIBROS XLSM/HISTORICO.xlsm", data_only=True)
        ws = wb["PRINCIPAL"]

        for row in ws.iter_rows(min_row=2, max_col=11):
            if escritura and row[1].value == escritura:
                self.line_edit_nir.setText(str(row[3].value))
                break
            elif nir and row[3].value == nir:
                self.line_edit_escritura.setText(str(row[1].value))
                break

def verificar_condiciones(row):
    columna_b = row[1].value
    columna_d = row[3].value
    columna_k = row[10].value
    return columna_b and not columna_k and columna_d != "No encontrado" and columna_d is not None

def consultar_nir(driver, nir, escritura_numero):
    try:
        input_nir = driver.find_element(By.ID, "filtroForm:j_idt41")
        input_nir.clear()
        input_nir.send_keys(nir)
        driver.find_element(By.ID, "filtroForm:j_idt43").click()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.ui-g-12.ui-md-6.ui-gl-6"))
        )

        container = driver.find_element(By.CSS_SELECTOR, "div.ui-g-12.ui-md-6.ui-gl-6")

        rechazado_texto = container.find_elements(By.XPATH, "//label[contains(text(), 'Documento en estado Ingreso Rechazado')]")
        
        if rechazado_texto:
            return f"Proceso Rechazado. Valide correcciones y envíe nuevamente."
        elif container.find_elements(By.XPATH, "//label[contains(text(), 'Se deben pagar todos los impuestos de registro para generar el recibo de pago')]") and container.find_elements(By.XPATH, "//label[contains(text(), 'Documento en estado Pago Pendiente')]"):
            return f"Aún no se ha cargado boleta de rentas para el caso."
        elif container.find_elements(By.XPATH, "//label[contains(text(), 'Se deben pagar todos los impuestos de registro para generar el recibo de pago')]") and container.find_elements(By.XPATH, "//label[contains(text(), 'Documento en estado Aprobación Pendiente')]"):
            return f"Caso en estado de aprobación PENDIENTE. Se debe cargar boleta de rentas al caso"
        elif container.find_elements    (By.XPATH, "//label[contains(text(), 'Se deben pagar todos los impuestos de registro para generar el recibo de pago')]") and container.find_elements(By.XPATH, "//label[contains(text(), 'Documento en estado Aprobación N/A')]"):
            return f"Caso en estado de aprobación 'N/A'. Se debe cargar boleta de rentas al caso"
        elif container.find_elements(By.XPATH, "//div[contains(text(), 'PAGAR EN LÍNEA')]"):
            return f"Recibo de pago descargado y sin cancelar"
        elif container.find_elements(By.XPATH, "//span[@class='ui-button-text ui-c' and contains(text(), 'Visualizar y generar')]"):
            return f"Recibo de pago listo para descargar"
        elif container.find_elements(By.XPATH, "//div[@style='font-size: 11px; font-weight: 600; color: #4bb04f' and contains(text(), 'PAGO REALIZADO')]"):
            return f"Recibo de pago descargado y cancelado"
        else:
            return f"No se puede determinar el estado del recibo"
    except TimeoutException:
        return f"Tiempo de espera agotado"
    except Exception as e:
        return f"Error durante la consulta: {str(e)}"

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    
    advanced_search_button = QPushButton("BUSQUEDA AVANZADA")
    advanced_search_button.clicked.connect(window.busqueda_avanzada)
    window.layout.addWidget(advanced_search_button)
    
    window.show()
    sys.exit(app.exec_())
