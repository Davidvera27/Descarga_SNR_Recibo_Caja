import os
import shutil
import sys
import openpyxl
from selenium.webdriver.common.keys import Keys
import time  # Importar la librería time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PyQt5 import QtWidgets
import win32com.client

class ConsultaAnexosInterfaz(QtWidgets.QWidget):
    def __init__(self, excel_file_path):
        super().__init__()
        self.setWindowTitle("Consulta de Anexos")
        self.setGeometry(100, 100, 800, 400)
        self.excel_file_path = excel_file_path
        self.proceso_en_curso = False

        # Cargar los estilos desde el archivo CSS
        with open("PYTHON\\WebScraping\\Excel\\Descarga_SNR_Recibo_Caja\\styles_Interfaz_Rentas.css", "r") as css_file:
            self.setStyleSheet(css_file.read())

        # Campos de búsqueda
        self.label_buscar_escritura = QtWidgets.QLabel("Buscar por ESCRITURA:", self)
        self.line_edit_escritura = QtWidgets.QLineEdit(self)
        self.label_buscar_escritura.setGeometry(50, 10, 150, 30)
        self.line_edit_escritura.setGeometry(200, 10, 150, 30)

        self.label_buscar_radicado = QtWidgets.QLabel("Buscar por RADICADO:", self)
        self.line_edit_radicado = QtWidgets.QLineEdit(self)
        self.label_buscar_radicado.setGeometry(400, 10, 150, 30)
        self.line_edit_radicado.setGeometry(550, 10, 150, 30)

        # Tabla de anexos
        self.tabla_anexos = QtWidgets.QTableWidget(self)
        self.tabla_anexos.setGeometry(50, 50, 700, 300)
        self.tabla_anexos.setColumnCount(6)  # Agregar dos columnas más
        self.tabla_anexos.setHorizontalHeaderLabels(["ESCRITURA", "RADICADO", "ESTADO DE LIQUIDACIÓN", "VER", "IMPRIMIR DOCUMENTO", "DESCARGAR DOCUMENTO"])
        self.tabla_anexos.verticalHeader().setVisible(False)
        self.tabla_anexos.horizontalHeader().setStretchLastSection(True)
        # Configurar el tamaño de la tabla después de crearla
        self.tabla_anexos.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)

        # Configurar el botón de consulta
        self.boton_consultar = QtWidgets.QPushButton("ABRIR LIQUIDACIÓN DE RENTAS", self)
        self.boton_consultar.setGeometry(50, 360, 250, 30)
        self.boton_consultar.clicked.connect(self.consultar_anexos)
        
        # Configurar el botón "VER TODO"
        self.boton_ver_todo = QtWidgets.QPushButton("VER TODO", self)
        self.boton_ver_todo.setGeometry(590, 360, 150, 30)
        self.boton_ver_todo.clicked.connect(self.ver_todos_los_anexos)

        # Configurar el botón de reiniciar
        self.boton_reiniciar = QtWidgets.QPushButton("REINICIAR", self)
        self.boton_reiniciar.setGeometry(320, 360, 250, 30)
        self.boton_reiniciar.clicked.connect(self.reiniciar_proceso)

        self.chrome_options = Options()
        self.chrome_options.add_argument("--headless")
        self.driver = webdriver.Chrome(options=self.chrome_options)

        # Segunda instancia del navegador para visualizar los anexos
        self.anexo_driver = None

        # Conectar los eventos textChanged de los QLineEdit a la función de filtrado
        self.line_edit_escritura.textChanged.connect(self.filtrar_tabla_por_escritura)
        self.line_edit_radicado.textChanged.connect(self.filtrar_tabla_por_radicado)

    def ver_anexo(self):
        boton = self.sender()
        fila = self.tabla_anexos.indexAt(boton.pos()).row()
        radicado = self.tabla_anexos.item(fila, 1).text()
        proceso = self.tabla_anexos.item(fila, 2).text()
        url_consulta = f"https://mercurio.antioquia.gov.co/mercurio/servlet/ControllerMercurio?command=anexos&tipoOperacion=abrirLista&idDocumento={radicado}&tipDocumento=R&now=Date()&ventanaEmergente=S&origen=NTR&proceso={proceso}"
        
        # Abrir el navegador controlado por Selenium en modo visible para mostrar los anexos
        self.anexo_driver = webdriver.Chrome()
        self.anexo_driver.get(url_consulta)

        try:
            # Esperar a que se cargue la tabla de anexos
            WebDriverWait(self.anexo_driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "tr.contenido td a[href*='EmergenteImagen']"))
            )
            # Encontrar el enlace dentro de la tabla y hacer clic en él
            enlace_anexo = self.anexo_driver.find_element(By.CSS_SELECTOR, "tr.contenido td a[href*='EmergenteImagen']")
            enlace_anexo.click()
        except Exception as e:
            print(f"Error al abrir el anexo: {e}")

    def imprimir_documento(self):
        # Implementar la funcionalidad de impresión del documento
        print("Imprimir documento")

    def descargar_documento(self):
        # Llamar a la función ver_anexo primero
        self.ver_anexo()

        # Esperar un momento para que se abra la tercera instancia del navegador
        time.sleep(5)  # Esto podría cambiarse por una espera más precisa usando las esperas de Selenium

        try:
            # Cambiar al contexto de la tercera ventana abierta
            self.anexo_driver.switch_to.window(self.anexo_driver.window_handles[-1])

            # Simular la combinación de teclas "CONTROL + S"
            shell = win32com.client.Dispatch("WScript.Shell")
            shell.SendKeys("^s")  # "^" indica la tecla CONTROL

            # Esperar un momento antes de enviar la tecla "ENTER"
            time.sleep(2)

            # Simular la tecla "ENTER" para guardar el archivo
            shell.SendKeys("{ENTER}")

            print("Archivo descargado correctamente.")
        except Exception as e:
            print(f"Error al descargar el archivo: {e}")

    def consultar_anexos(self):
        
        radicados_a_consultar = self.buscar_nir_para_consulta(self.excel_file_path)
        self.proceso_en_curso = True

        for radicado, escritura_procesada in radicados_a_consultar:
            try:
                estado_liquidacion, enlaces_anexos = self.obtener_estado_y_enlaces_anexos(radicado)
                if enlaces_anexos:
                    self.agregar_fila_tabla_anexos(escritura_procesada, radicado, estado_liquidacion)
            except Exception as e:
                print(f"Error al consultar radicado {radicado}: {e}")

        print("Proceso de consulta de anexos finalizado.")
        self.proceso_en_curso = False

    def buscar_nir_para_consulta(self, excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheet = wb["ESCRITURAS FISICAS (DIARIO)"]

        # Buscar la última fila con datos en la columna D (columna 4)
        last_row = sheet.max_row
        for row in range(last_row, 1, -1):
            if sheet.cell(row=row, column=4).value == "SIN INICIAR":
                last_valid_row = row
                break

        # Obtener los datos de las filas relevantes
        rows = sheet.iter_rows(min_row=2, max_row=last_valid_row, max_col=4, values_only=True)
        
        return [(row[2], row[1]) for row in rows if row[3] == "SIN INICIAR" and row[1]]

    def obtener_estado_y_enlaces_anexos(self, radicado):
        url_consulta = f"https://mercurio.antioquia.gov.co/mercurio/servlet/ControllerMercurio?command=anexos&tipoOperacion=abrirLista&idDocumento={radicado}&tipDocumento=R&now=Date()&ventanaEmergente=S&origen=NTR"
        self.driver.get(url_consulta)
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "//td[contains(text(), 'Total Anexos')]")))
        enlaces_anexos = self.driver.find_elements(By.XPATH, "//tr[@class='contenido']/td/a")
        cantidad_anexos = len(enlaces_anexos) if enlaces_anexos else 0
        estado_liquidacion = f"LIQUIDACIÓN LISTA ({cantidad_anexos})" if cantidad_anexos else "SIN ANEXOS"
        return estado_liquidacion, enlaces_anexos

    def agregar_fila_tabla_anexos(self, escritura_procesada, radicado, estado_liquidacion):
        fila = self.tabla_anexos.rowCount()
        self.tabla_anexos.insertRow(fila)
        self.tabla_anexos.setItem(fila, 0, QtWidgets.QTableWidgetItem(str(escritura_procesada)))
        self.tabla_anexos.setItem(fila, 1, QtWidgets.QTableWidgetItem(str(radicado)))
        self.tabla_anexos.setItem(fila, 2, QtWidgets.QTableWidgetItem(estado_liquidacion))

        # Configurar los botones de las nuevas columnas
        self.configurar_boton(fila, 3, "VER", self.ver_anexo)
        self.configurar_boton(fila, 4, "IMPRIMIR", self.imprimir_documento)
        self.configurar_boton(fila, 5, "DESCARGAR", self.descargar_documento)

    def configurar_boton(self, fila, columna, texto, funcion):
        boton = QtWidgets.QPushButton(texto)
        boton.clicked.connect(funcion)
        self.tabla_anexos.setCellWidget(fila, columna, boton)
        
    def ver_todos_los_anexos(self):
        # Define radicados here or obtain it from somewhere else
        radicados = [...]  # Define your list of radicados here

        # Inicializar self.anexo_driver si aún no se ha hecho
        if self.anexo_driver is None:
            self.anexo_driver = webdriver.Chrome()

        # Procesar cada radicado para abrir los anexos
        for radicado in radicados:
            try:
                estado_liquidacion, enlaces_anexos = self.obtener_estado_y_enlaces_anexos(radicado)
                if enlaces_anexos:
                    self.abrir_todos_los_enlaces(enlaces_anexos)
            except Exception as e:
                print(f"Error al consultar radicado {radicado}: {e}")

        print("Proceso de apertura de todos los anexos finalizado.")

    def abrir_todos_los_enlaces(self, enlaces_anexos):
        # Abrir cada enlace de anexo uno por uno
        for enlace in enlaces_anexos:
            enlace.click()
            # Esperar un tiempo antes de abrir el siguiente enlace
            time.sleep(2)  # Se puede ajustar según sea necesario

            # Cambiar al contexto de la tercera ventana abierta
            self.anexo_driver.switch_to.window(self.anexo_driver.window_handles[-1])

            # Simular la combinación de teclas "CONTROL + S" para descargar el archivo
            shell = win32com.client.Dispatch("WScript.Shell")
            shell.SendKeys("^s")  # "^" indica la tecla CONTROL

            # Esperar un momento antes de enviar la tecla "ENTER"
            time.sleep(2)
            shell.SendKeys("{ENTER}")

            print("Archivo descargado correctamente.")

    def reiniciar_proceso(self):
        self.tabla_anexos.clearContents()
        self.tabla_anexos.setRowCount(0)
        self.consultar_anexos()

    def filtrar_tabla_por_escritura(self, texto):
        self.filtrar_tabla_por_columna(texto, 0)

    def filtrar_tabla_por_radicado(self, texto):
        self.filtrar_tabla_por_columna(texto, 1)

    def filtrar_tabla_por_columna(self, texto, columna):
        for fila in range(self.tabla_anexos.rowCount()):
            item = self.tabla_anexos.item(fila, columna)
            if item:
                self.tabla_anexos.setRowHidden(fila, texto.lower() not in item.text().lower())

if __name__ == "__main__":
    excel_file_path = r'C:\\Users\\DAVID\\Desktop\\DAVID\\N-15\\DAVID\\LIBROS XLSM\\HISTORICO.xlsm'
    app = QtWidgets.QApplication(sys.argv)
    ventana = ConsultaAnexosInterfaz(excel_file_path)
    ventana.show()
    sys.exit(app.exec_())
