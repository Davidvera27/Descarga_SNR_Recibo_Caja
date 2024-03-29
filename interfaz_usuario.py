import tkinter as tk
from tkinter import messagebox, ttk
from threading import Thread
import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os

import Descargar_Rentas as descargar_rentas
from inicio_sesion import iniciar_sesion
from buscar_documentos import buscar_documentos
from descargar_archivos import descargar_archivo

def mostrar_detalles(resultados, indice):
    detalle_ventana = tk.Toplevel()
    detalle_ventana.title("Detalles del Resultado")

    detalle_frame = ttk.Frame(detalle_ventana)
    detalle_frame.pack(padx=10, pady=10)

    for i, detalle in enumerate(resultados[indice]):
        ttk.Label(detalle_frame, text=f"Columna {i+1}:").grid(row=i, column=0, sticky="e", padx=5, pady=5)
        ttk.Label(detalle_frame, text=detalle).grid(row=i, column=1, sticky="w", padx=5, pady=5)

def mostrar_resultados(resultados):
    ventana_resultados = tk.Toplevel()
    ventana_resultados.title("Resultados de la consulta")

    tabla_frame = ttk.Frame(ventana_resultados)
    tabla_frame.pack(expand=True, fill="both")
    tabla = ttk.Treeview(tabla_frame, columns=("#1", "#2", "#3", "#4", "#5", "#6", "#7", "#8", "#9", "#10", "#11"))
    tabla.heading("#0", text="Acciones")
    tabla.heading("#1", text="Ver")
    tabla.heading("#2", text="Columna Disponible")
    tabla.heading("#3", text="Escritura")
    tabla.heading("#4", text="Estado del proceso")
    tabla.heading("#5", text="NIR")
    tabla.heading("#6", text="Vigencia de Rentas")
    tabla.heading("#7", text="Estado en el REL")
    tabla.heading("#8", text="Tramitadora")
    tabla.heading("#9", text="Número de paquete")
    tabla.heading("#10", text="Estado de liquidación")
    tabla.heading("#11", text="Recibo de pago (Estado)")

    tabla.column("#0", width=100)

    scrollbar_x = ttk.Scrollbar(tabla_frame, orient="horizontal", command=tabla.xview)
    tabla.configure(xscrollcommand=scrollbar_x.set)
    scrollbar_x.pack(side="bottom", fill="x")

    for i, resultado in enumerate(resultados):
        tabla.insert("", "end", iid=i, values=(f"Ver Detalles {i}", *resultado))

    tabla.pack(expand=True, fill="both")
    tabla.bind("<Double-1>", lambda event: mostrar_detalles(resultados, int(tabla.selection()[0])))

def exportar_a_excel(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active

    for fila in resultados:
        ws.append(fila)

    wb.save("resultados.xlsx")

def mostrar_mensaje(mensaje):
    messagebox.showinfo("Mensaje", mensaje)

def guardar_consulta_en_historial(consulta, resultados):
    with open("historial.txt", "a") as archivo:
        archivo.write(f"Consulta: {consulta}\n")
        for resultado in resultados:
            archivo.write(f"{resultado}\n")
        archivo.write("\n")

def consultar_anexos():
    service = Service('PYTHON\\WebScraping\\chromedriver-win64\\chromedriver.exe')
    options = Options()
    options.add_argument('--start-maximized')
    driver = WebDriver(service=service, options=options)

    try:
        usuario = usuario_entry.get()
        contraseña = contraseña_entry.get()
        iniciar_sesion(driver, usuario, contraseña)

        archivo_excel = "C:/Users/DAVID/Desktop/DAVID/N-15/DAVID/LIBROS XLSM/HISTORICO.xlsm"
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        sheet = wb["PRINCIPAL"]

        resultados = []

        for row in sheet.iter_rows(min_row=2, max_col=11, values_only=False):
            if row[10].value == "DESCARGADA":
                nir = row[3].value
                if nir is not None:
                    buscar_documentos(driver, nir)
                    resultado_descarga = descargar_archivo(driver, row)
                    resultados.append([cell.value for cell in row])
                    resultados[-1].append(resultado_descarga)
                else:
                    print(f"La escritura {row[1].value} no contiene NIR")

        mostrar_resultados(resultados)
        exportar_a_excel(resultados)
        mostrar_mensaje("Proceso de consulta completado")
        guardar_consulta_en_historial("Consulta de anexos", resultados)

    except Exception as e:
        mostrar_mensaje(f"Error: {e}")
    finally:
        driver.quit()

def consultar_recibo_pago():
    os.system('start /wait cmd /c "python PYTHON\WebScraping\Excel\Descarga_SNR_Recibo_Caja\Descargar_ReciboPago.py"')

def consultar_rentas():
    descargar_rentas.consultar_anexos()

root = tk.Tk()
root.title("Interfaz de Usuario")

tk.Label(root, text="Usuario:").grid(row=0, column=0, padx=5, pady=5)
usuario_entry = tk.Entry(root)
usuario_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Label(root, text="Contraseña:").grid(row=1, column=0, padx=5, pady=5)
contraseña_entry = tk.Entry(root, show="*")
contraseña_entry.grid(row=1, column=1, padx=5, pady=5)

def iniciar_proceso():
    Thread(target=consultar_anexos).start()
tk.Button(root, text="Descargar Recibo de caja", command=iniciar_proceso).grid(row=2, columnspan=2, padx=5, pady=5)

tk.Button(root, text="Consultar Recibo de Pago", command=consultar_recibo_pago).grid(row=3, columnspan=2, padx=5, pady=5)

def proceso4():
    consultar_rentas()
tk.Button(root, text="Consultar Liquidación de Rentas", command=proceso4).grid(row=4, columnspan=2, padx=5, pady=5)

root.mainloop()